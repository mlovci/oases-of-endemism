import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_squared_error
from sklearn.preprocessing import StandardScaler
import statsmodels.api as sm
from joblib import Parallel, delayed

# Ensure figures directory exists
os.makedirs("figures", exist_ok=True)

# Set publication quality styling
sns.set_theme(style="ticks", context="talk")
plt.rcParams.update({
    "font.family": "sans-serif",
    "font.sans-serif": ["Arial", "Calibri", "DejaVu Sans"],
    "figure.titlesize": 16,
    "axes.titlesize": 12,
    "axes.labelsize": 11,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "savefig.bbox": "tight"
})

def run_single_bootstrap(X_scaled, y, seed):
    np.random.seed(seed)
    train_idx = np.random.choice(X_scaled.index, size=len(X_scaled), replace=True)
    val_idx = np.array(list(set(X_scaled.index) - set(train_idx)))
    
    X_train, y_train = X_scaled.loc[train_idx], y.loc[train_idx]
    X_val, y_val = X_scaled.loc[val_idx], y.loc[val_idx]
    
    model = RandomForestRegressor(
        n_estimators=100,
        max_depth=4,
        max_features="sqrt",
        random_state=seed
    )
    model.fit(X_train, y_train)
    
    y_pred = model.predict(X_val)
    try:
        val_r2 = r2_score(y_val, y_pred)
    except Exception:
        val_r2 = np.nan
    val_mse = mean_squared_error(y_val, y_pred)
    
    return model.feature_importances_, val_r2, val_mse

def main():
    print("Loading dataset...")
    df = pd.read_excel("Data_Table_S5.xlsx", header=[0, 1])
    df.columns = df.columns.get_level_values(0)
    
    # Filter to Regional Aquifer springs (N=45)
    reg = df[df["Aquifer Type"] == "Regional Aq"].copy()
    
    # Target biological variables
    taxa = ["Endemics", "Crenophilies", "Springsnails", "Non Natives", "Native Fish"]
    taxa_labels = {
        "Endemics": "Endemic Richness",
        "Crenophilies": "Crenophile Richness",
        "Springsnails": "Springsnail Richness",
        "Non Natives": "Non-Native Richness",
        "Native Fish": "Native Fish Richness"
    }
    
    # Define physical environmental features (15 variables)
    features_15 = ["Depth", "Width", "Temperature", "Conductivity", "pH", 
                   "Emerge Cover", "Bank Cover", "Silt", "Sand", "Gravel", "Cobble",
                   "Diversion", "Equine", "Cattle", "Recreate"]
                   
    # Define all potential features including Non Natives (16 variables)
    features_16 = features_15 + ["Non Natives"]
    
    n_iterations = 1000
    results_summary = {}
    importances_dict = {}
    glm_coefficients = {}
    
    fig, axes = plt.subplots(3, 2, figsize=(15, 18))
    axes = axes.flatten()
    
    # Color palette for plots
    colors = sns.color_palette("muted", len(taxa))
    
    for idx, taxon in enumerate(taxa):
        print(f"\nAnalyzing target taxon: {taxon}...")
        y = reg[taxon].astype(float)
        
        # Determine features list for this taxon: exclude Non Natives only for Non Natives itself
        feats = features_15 if taxon == "Non Natives" else features_16
        
        X_raw = reg[feats].astype(float)
        scaler = StandardScaler()
        X_scaled = pd.DataFrame(scaler.fit_transform(X_raw), columns=feats, index=reg.index)
        
        # 1. Run Parallel Bootstrapping
        results = Parallel(n_jobs=-1)(
            delayed(run_single_bootstrap)(X_scaled, y, 100 + i) for i in range(n_iterations)
        )
        
        importances_list = [r[0] for r in results]
        r2_scores = [r[1] for r in results if not np.isnan(r[1])]
        mse_scores = [r[2] for r in results]
        
        mean_r2 = np.mean(r2_scores)
        median_r2 = np.median(r2_scores)
        mean_mse = np.mean(mse_scores)
        
        importances_df = pd.DataFrame(importances_list, columns=feats)
        mean_importances = importances_df.mean()
        # Pad with 0.0 or nan for non-applicable features to keep consistent dimensions
        for f in features_16:
            if f not in mean_importances:
                mean_importances[f] = np.nan
                
        importances_dict[taxon] = mean_importances
        
        print(f"  OOS Validation: R2={mean_r2:.4f} (median={median_r2:.4f}), MSE={mean_mse:.4f}")
        print("  Top 5 predictors:")
        sorted_imp = mean_importances.dropna().sort_values(ascending=False)
        for f, val in sorted_imp.head(5).items():
            print(f"    - {f:12}: {val:.4f}")
            
        results_summary[taxon] = {
            "Mean R2": mean_r2,
            "Median R2": median_r2,
            "Mean MSE": mean_mse,
            "Top Predictor": sorted_imp.index[0],
            "Top Predictor Importance": sorted_imp.values[0]
        }
        
        # 2. Fit Standardized Poisson GLM for regression coefficients
        X_const = sm.add_constant(X_scaled)
        try:
            glm = sm.GLM(y, X_const, family=sm.families.Poisson()).fit(cov_type='HC3')
            coefs = glm.params
            pvalues = glm.pvalues
            glm_coefficients[taxon] = {}
            for f in ["const"] + features_16:
                if f in X_const.columns:
                    glm_coefficients[taxon][f] = (coefs[f], pvalues[f])
                else:
                    glm_coefficients[taxon][f] = (np.nan, np.nan)
        except Exception as e:
            print(f"  GLM fitting failed: {e}")
            glm_coefficients[taxon] = {
                f: (np.nan, np.nan) for f in ["const"] + features_16
            }
            
        # 3. Plot Feature Importance panel
        ax = axes[idx]
        plot_imp = mean_importances.dropna().sort_values(ascending=True)
        ax.barh(plot_imp.index, plot_imp.values, color=colors[idx], edgecolor='gray', height=0.6)
        ax.set_title(f"Panel {chr(65+idx)}: {taxa_labels[taxon]} ($R^2_{{median}} = {median_r2:.3f}$)", fontsize=13, fontweight='bold')
        ax.set_xlabel("Mean Gini Importance")
        ax.set_xlim(0, 0.35)
        ax.grid(axis='x', linestyle='--', alpha=0.5)
        ax.tick_params(labelsize=9)
        
    # Hide the 6th (empty) subplot
    axes[5].axis('off')
    # Add a summary legend/text inside the 6th subplot
    text_info = (
        "Statistical Summary (Regional Springs, N=45):\n\n"
        "• Features are normalized via standard scaling.\n"
        "• Random Forests fit with 100 trees, max depth = 4,\n"
        "  max features = 'sqrt'.\n"
        "• Metrics calculated over 1,000 bootstrap splits.\n"
        "• Out-of-Sample validation uses out-of-bag (OOB)\n"
        "  unseen validation folds.\n\n"
        "Figures saved as print-ready PDF and PNG."
    )
    axes[5].text(0.05, 0.2, text_info, fontsize=10, style='italic', bbox=dict(boxstyle='round', facecolor='whitesmoke', alpha=0.8))
    
    plt.tight_layout()
    plt.savefig("figures/Figure_S6_Taxa_Feature_Importances.png", dpi=300)
    plt.savefig("figures/Figure_S6_Taxa_Feature_Importances.pdf", dpi=300)
    plt.close()
    print("\nSaved multi-panel feature importance plot: figures/Figure_S6_Taxa_Feature_Importances.png and .pdf")
    
    # 4. Write Poisson GLM coefficients table
    print("\n" + "="*80)
    print("STANDARD REGRESSION ANALYSIS (Standardized Poisson GLM HC3 Coefficients)")
    print("="*80)
    
    glm_df_data = []
    for f in ["const"] + features_16:
        row = {"Feature": f}
        for taxon in taxa:
            val, p = glm_coefficients[taxon][f]
            if pd.isna(val) or pd.isna(p):
                row[taxa_labels[taxon]] = "N/A"
            else:
                sig = ""
                if p < 0.001: sig = "***"
                elif p < 0.01: sig = "**"
                elif p < 0.05: sig = "*"
                row[taxa_labels[taxon]] = f"{val:+.4f}{sig} (p={p:.3e})"
        glm_df_data.append(row)
        
    # Manually print the markdown table to avoid requiring 'tabulate' dependency
    cols = ["Feature"] + [taxa_labels[t] for t in taxa]
    print("| " + " | ".join(cols) + " |")
    print("|" + "|".join(["---" for _ in cols]) + "|")
    for r in glm_df_data:
        vals = [r["Feature"]] + [r[taxa_labels[t]] for t in taxa]
        print("| " + " | ".join(vals) + " |")
    
    # 5. Export to Excel Table 6
    print("\nExporting regression and feature importance results to Table_6_Taxa_Regression.xlsx...")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Style definitions
    font_title = Font(name="Calibri", size=13, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="000000")
    font_body = Font(name="Calibri", size=11, color="000000")
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Sheet 1: GLM Coefficients
    ws1 = wb.active
    ws1.title = "Poisson GLM Coefficients"
    ws1.views.sheetView[0].showGridLines = True
    ws1.append(["Table 6a: Standardized Poisson GLM HC3 Coefficients in Regional Aquifer Springs (N=45)"])
    ws1.cell(1, 1).font = font_title
    ws1.append([])
    
    headers1 = ["Feature"] + [taxa_labels[t] for t in taxa]
    ws1.append(headers1)
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(3, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, r in enumerate(glm_df_data, 4):
        vals = [r["Feature"]] + [r[taxa_labels[t]] for t in taxa]
        ws1.append(vals)
        for col_idx in range(1, len(vals) + 1):
            cell = ws1.cell(row_idx, col_idx)
            cell.font = font_body
            cell.border = border_thin
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Sheet 2: RF Feature Importances
    ws2 = wb.create_sheet(title="Random Forest Importances")
    ws2.views.sheetView[0].showGridLines = True
    ws2.append(["Table 6b: Mean Random Forest Feature Importances across 1,000 Bootstrap Splits"])
    ws2.cell(1, 1).font = font_title
    ws2.append([])
    
    headers2 = ["Feature"] + [taxa_labels[t] for t in taxa]
    ws2.append(headers2)
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(3, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, f in enumerate(features_16, 4):
        vals = [f] + [importances_dict[taxon][f] for taxon in taxa]
        ws2.append(vals)
        for col_idx in range(1, len(vals) + 1):
            cell = ws2.cell(row_idx, col_idx)
            cell.font = font_body
            cell.border = border_thin
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                if not pd.isna(cell.value):
                    cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-width
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Sheet 3: Performance Metrics
    ws3 = wb.create_sheet(title="Model Validation Performance")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["Table 6c: Random Forest Regressor Out-of-Sample Validation Performance (1,000 Splits)"])
    ws3.cell(1, 1).font = font_title
    ws3.append([])
    
    headers3 = ["Taxon Metric", "Mean OOS R2", "Median OOS R2", "Mean OOS MSE", "Top Predictor"]
    ws3.append(headers3)
    for col_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(3, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, taxon in enumerate(taxa, 4):
        summary = results_summary[taxon]
        vals = [taxa_labels[taxon], summary["Mean R2"], summary["Median R2"], summary["Mean MSE"], summary["Top Predictor"]]
        ws3.append(vals)
        for col_idx in range(1, len(vals) + 1):
            cell = ws3.cell(row_idx, col_idx)
            cell.font = font_body
            cell.border = border_thin
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
            if col_idx == 1 or col_idx == 5:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-width
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save("Table_6_Taxa_Regression.xlsx")
    print("Successfully exported Table_6_Taxa_Regression.xlsx with sheets for GLM, RF, and Validation Performance.")

if __name__ == "__main__":
    main()
